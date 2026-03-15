import sys
import os
import pandas as pd
import json
import glob
from PySide6.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QListWidget, QLabel, QFileDialog, QMessageBox, QComboBox,
                             QProgressBar, QTextEdit, QTableWidget, QTableWidgetItem, QHeaderView)
from PySide6.QtCore import Qt, QThread, Signal
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

# === CLINE MODIFIED START ===
# 原有逻辑参考：
# 之前的实现是在主线程中直接进行文件合并操作，会导致UI卡顿。
# 重构为使用 QThread，将合并任务放到后台线程执行，通过信号更新UI进度。

# --- 1. 合并执行线程 (使用 QThread 重构) ---
# === CLINE MODIFIED START - MergeWorker 类 ===
class MergeWorker(QThread):
    progress = Signal(int)
    log = Signal(str)
    task_finished = Signal(pd.DataFrame)

    def __init__(self, file_paths, template_config):
        super().__init__()
        self.file_paths = file_paths
        self.template_config = template_config

    def run(self):
        try:
            target_headers = self.template_config["headers"]
            all_dfs = []
            
            for i, path in enumerate(self.file_paths):
                self.log.emit(f"正在处理: {os.path.basename(path)}")
                df = pd.read_csv(path) if path.endswith('.csv') else pd.read_excel(path)
                
                current_headers = df.columns.tolist()
                if not set(target_headers).issubset(set(current_headers)):
                    missing_cols = set(target_headers) - set(current_headers)
                    self.log.emit(f"❌ 错误: {os.path.basename(path)} 缺少列: {missing_cols}")
                    continue
                
                df = df[target_headers]
                for col, settings in self.template_config["column_settings"].items():
                    if settings["null_policy"] == "删除空值行":
                        df = df.dropna(subset=[col])
                    if settings["duplicate_policy"] == "仅留首行":
                        df = df.drop_duplicates(subset=[col], keep='first')

                all_dfs.append(df)
                self.progress.emit(int((i + 1) / len(self.file_paths) * 100))

            if not all_dfs:
                self.log.emit("未发现有效匹配文件。")
                return

            final_df = pd.concat(all_dfs, ignore_index=True)
            self.log.emit(f"✅ 合并完成，共计 {len(final_df)} 行。")
            self.task_finished.emit(final_df)

        except Exception as e:
            self.log.emit(f"🔥 错误: {str(e)}")
# === CLINE MODIFIED END - MergeWorker 类 ===

# === CLINE MODIFIED START ===
# 原有逻辑参考：
# start_merge 方法之前直接在主线程中执行合并操作，会导致UI阻塞。
# 重构为创建 MergeWorker 实例并在后台线程运行，通过信号槽机制更新进度。

# --- 2. 模块 C 主界面 ---
class MultiFileMergePage(QWidget):
    def __init__(self):
        super().__init__()
        self.template_config = None
        self.merged_df = None
        
        self.setStyleSheet("""
            QWidget { background-color: #2F3542; color: #F1F2F6; font-family: 'Segoe UI', 'Microsoft YaHei'; }
            QLabel { color: #A4B0BE; font-weight: bold; }
            QListWidget { background-color: #353B48; border: 1px solid #475062; border-radius: 4px; padding: 5px; }
            QTextEdit { background-color: #1E2229; color: #2ecc71; border: 1px solid #475062; font-family: 'Consolas'; }
            QTableWidget { background-color: #353B48; gridline-color: #475062; border: 1px solid #475062; color: #F1F2F6; }
            QHeaderView::section { background-color: #57606F; color: white; border: 1px solid #475062; }
            QProgressBar { border: 1px solid #475062; border-radius: 4px; text-align: center; background: #353B48; }
            QProgressBar::chunk { background-color: #70A1FF; }
            QComboBox { background-color: #353B48; border: 1px solid #70A1FF; border-radius: 4px; padding: 5px; color: white; }
            QComboBox::drop-down { border: none; }
            QPushButton { background-color: #57606F; border: 1px solid #747D8C; padding: 8px; border-radius: 4px; color: white; }
            QPushButton:hover { background-color: #747D8C; }
        """)
        
        self.init_ui()
        self.refresh_template_list() # 初始化时扫描

    def init_ui(self):
        self.setFixedSize(1600, 800)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(25)

        # --- 左侧：配置面板 ---
        left_panel = QVBoxLayout()
        
        # 模板选择区域 (新功能)
        left_panel.addWidget(QLabel("⚙️ 配置模板选择:"))
        tpl_layout = QHBoxLayout()
        self.combo_templates = QComboBox()
        self.btn_refresh_tpl = QPushButton("🔄")
        self.btn_refresh_tpl.setFixedWidth(40)
        tpl_layout.addWidget(self.combo_templates)
        tpl_layout.addWidget(self.btn_refresh_tpl)
        left_panel.addLayout(tpl_layout)
        
        self.label_template_status = QLabel("🔴 未载入配置")
        left_panel.addWidget(self.label_template_status)
        
        left_panel.addSpacing(15)
        left_panel.addWidget(QLabel("待处理文件队列:"))
        self.file_list = QListWidget()
        left_panel.addWidget(self.file_list)
        
        self.btn_select = QPushButton("📂 添加待合并数据")
        self.btn_run = QPushButton("🚀 执行合并任务")
        self.btn_run.setEnabled(False)
        self.btn_run.setStyleSheet("background-color: #2ecc71; color: white; font-weight: bold; height: 45px;")
        
        left_panel.addWidget(self.btn_select)
        left_panel.addWidget(self.btn_run)
        
        # --- 右侧：状态与展示 ---
        right_panel = QVBoxLayout()
        right_panel.addWidget(QLabel("系统运行日志:"))
        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        right_panel.addWidget(self.log_view)
        
        right_panel.addSpacing(15)
        right_panel.addWidget(QLabel("数据预览 (Top 5):"))
        self.preview_table = QTableWidget()
        self.preview_table.setFixedHeight(220)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        right_panel.addWidget(self.preview_table)
        
        self.pbar = QProgressBar()
        right_panel.addWidget(self.pbar)
        
        self.btn_export = QPushButton("💾 导出并应用样式 (.xlsx)")
        self.btn_export.setEnabled(False)
        self.btn_export.setStyleSheet("background-color: #70A1FF; color: white; font-weight: bold; height: 45px;")
        right_panel.addWidget(self.btn_export)

        layout.addLayout(left_panel, 1)
        layout.addLayout(right_panel, 2)

        # 信号绑定
        self.btn_refresh_tpl.clicked.connect(self.refresh_template_list)
        self.combo_templates.currentIndexChanged.connect(self.load_selected_template)
        self.btn_select.clicked.connect(self.select_files)
        self.btn_run.clicked.connect(self.start_merge)
        self.btn_export.clicked.connect(self.export_data)

    def refresh_template_list(self):
        """扫描当前目录下的所有 JSON 模板文件"""
        self.combo_templates.clear()
        json_files = glob.glob("*.json")
        if not json_files:
            self.label_template_status.setText("🔴 目录下未发现 .json 模板")
            self.btn_run.setEnabled(False)
            return
        
        self.combo_templates.addItems(json_files)
        # 如果有名为“模板.json”的，默认选中它
        if "模板.json" in json_files:
            self.combo_templates.setCurrentText("模板.json")
        self.load_selected_template()

    def load_selected_template(self):
        """根据下拉框加载配置"""
        file_name = self.combo_templates.currentText()
        if not file_name: return
        
        try:
            with open(file_name, "r", encoding="utf-8") as f:
                self.template_config = json.load(f)
            self.label_template_status.setText(f"🟢 已载入: {file_name}")
            self.label_template_status.setStyleSheet("color: #2ecc71; font-weight: bold;")
            self.btn_run.setEnabled(True)
        except Exception as e:
            self.label_template_status.setText("❌ 配置读取失败")
            self.btn_run.setEnabled(False)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "选择数据文件", "", "Excel/CSV (*.xlsx *.csv)")
        if files:
            self.file_list.clear()
            self.file_list.addItems(files)

    def start_merge(self):
        if not self.file_list.count(): 
            QMessageBox.warning(self, "提醒", "请先添加待处理文件！")
            return
        paths = [self.file_list.item(i).text() for i in range(self.file_list.count())]
        self.log_view.clear()
        self.pbar.setValue(0)
        self.worker = MergeWorker(paths, self.template_config)
        self.worker.progress.connect(self.pbar.setValue)
        self.worker.log.connect(self.log_view.append)
        self.worker.task_finished.connect(self.on_merge_finished)
        self.worker.start()

    def on_merge_finished(self, df):
        self.merged_df = df
        self.btn_export.setEnabled(True)
        
        self.preview_table.setColumnCount(len(df.columns))
        self.preview_table.setRowCount(min(5, len(df)))
        self.preview_table.setHorizontalHeaderLabels(df.columns)
        for r in range(min(5, len(df))):
            for c in range(len(df.columns)):
                self.preview_table.setItem(r, c, QTableWidgetItem(str(df.iloc[r, c])))
        
        self.log_view.append("\n✅ 合并清洗任务圆满完成。")

    def export_data(self):
        save_path, _ = QFileDialog.getSaveFileName(self, "存储路径", "merged_output.xlsx", "Excel Files (*.xlsx)")
        if not save_path: return

        try:
            self.log_view.append(f"正在导出...")
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                self.merged_df.to_excel(writer, index=False, sheet_name='Data')
                ws = writer.sheets['Data']
                
                font_std = Font(name='等线', size=11)
                font_bold = Font(name='等线', size=11, bold=True)
                no_border = Border(left=Side(style=None), right=Side(style=None), 
                                 top=Side(style=None), bottom=Side(style=None))
                
                for i, col in enumerate(self.merged_df.columns):
                    col_letter = get_column_letter(i + 1)
                    max_len = max(self.merged_df[col].astype(str).map(len).max(), len(str(col))) + 2
                    ws.column_dimensions[col_letter].width = min(max_len, 50)
                    
                    for row in range(1, len(self.merged_df) + 2):
                        cell = ws.cell(row=row, column=i+1)
                        cell.font = font_bold if row == 1 else font_std
                        cell.border = no_border

            self.log_view.append("🎉 导出成功！")
            QMessageBox.information(self, "通知", "文件已按“等线”字体、无边框格式导出。")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {e}")

# === CLINE MODIFIED END ===

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MultiFileMergePage()
    window.show()
    sys.exit(app.exec())
